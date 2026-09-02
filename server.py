from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
import json, os, hashlib, sqlite3, io, datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from dotenv import load_dotenv
import anthropic

load_dotenv()

app = Flask(__name__)

# ASVS Level-weighted severity scoring. Since ASVS 5.0.0 discontinued CWE mappings,
# we use ASVS Level as the severity signal instead: Level 1 = baseline hygiene every
# app should have (weight 3), Level 2 = most business apps (weight 2), Level 3 =
# high-assurance/advanced controls (weight 1). Missing a Level 1 control counts more
# against the score than missing a Level 3 one.
#
# Maturity level classification follows how ASVS actually works: levels are cumulative,
# not a single blended percentage. "Level 2" requires near-complete Level 1 coverage
# PLUS solid combined L1+L2 coverage — you cannot claim Level 2 while missing basic
# Level 1 controls.
LEVEL_WEIGHTS = {"1": 3, "2": 2, "3": 1}

def compute_weighted_score(reqs):
    if not reqs:
        return 0, "Not Assessed"
    total_weight = earned_weight = 0
    l1_total = l1_done = l12_total = l12_done = 0
    for r in reqs:
        lvl = str(r.get("level", "1"))
        w = LEVEL_WEIGHTS.get(lvl, 1)
        done = bool(r.get("implemented"))
        total_weight += w
        if done: earned_weight += w
        if lvl == "1":
            l1_total += 1
            if done: l1_done += 1
        if lvl in ("1", "2"):
            l12_total += 1
            if done: l12_done += 1
    pct = round((earned_weight/total_weight)*100) if total_weight else 0
    l1_pct = (l1_done/l1_total*100) if l1_total else 100
    l12_pct = (l12_done/l12_total*100) if l12_total else 100
    if l1_pct >= 90 and l12_pct >= 70:
        level = "Level 2"
    elif l1_pct >= 70:
        level = "Level 1"
    else:
        level = "Below Level 1"
    return pct, level
CORS(app)

ANTHROPIC_MODEL = "claude-opus-5"
USE_CLAUDE_AI = os.getenv("USE_CLAUDE_AI", "false").lower() == "true"
AI_MAX_CODE_CHARS = 120000  # keep the prompt well within the context window

_anthropic_client = None
def get_anthropic_client():
    global _anthropic_client
    if _anthropic_client is None:
        api_key = os.getenv("ANTHROPIC_API_KEY")
        if not api_key:
            return None
        _anthropic_client = anthropic.Anthropic(api_key=api_key)
    return _anthropic_client

# ══════════════════════════════════════════════════════════════════════════
#  CLAUDE AI SECURITY ASSESSMENT — OUTPUT CONTRACT
#  Model: claude-opus-5   |   effort: "low"  (fast, ~10s, keeps replies short)
#  Enforced on the API call via output_config.format (JSON schema) — Claude
#  cannot return anything outside this shape:
#
#    securityScore        int, 0-100          Claude's own judgment
#    summary              string              1-2 short, plain-language sentences
#    strengths            array, max 3 items  one plain sentence each
#    risks                array, max 3 items  one plain sentence each
#    topRecommendations   array, max 3 items  one actionable step each
# ══════════════════════════════════════════════════════════════════════════
AI_INSIGHTS_SCHEMA = {
    "type": "object",
    "properties": {
        "securityScore": {"type": "integer", "description": "Overall security posture, 0-100, based on your own review of the code"},
        "summary": {"type": "string", "description": "One or two short, plain-language sentences on the codebase's overall security posture. No jargon, no code citations."},
        "strengths": {"type": "array", "items": {"type": "string"}, "description": "At most 3 items. Each is one short, plain sentence naming a real security control found in the code."},
        "risks": {"type": "array", "items": {"type": "string"}, "description": "At most 3 items. Each is one short, plain sentence naming a concrete risk or missing control, in everyday language."},
        "topRecommendations": {"type": "array", "items": {"type": "string"}, "description": "At most 3 items. Each is one short, actionable next step a developer can take."}
    },
    "required": ["securityScore", "summary", "strengths", "risks", "topRecommendations"],
    "additionalProperties": False
}

@app.route('/api/ai/analyze', methods=['POST'])
def ai_analyze():
    if not USE_CLAUDE_AI:
        return jsonify({
            "ok": True,
            "securityScore": 0,
            "summary": "AI analysis is disabled during development (USE_CLAUDE_AI=false). Enable it before final testing/demo.",
            "strengths": [],
            "risks": [],
            "topRecommendations": [],
            "aiDisabled": True
        })

    client = get_anthropic_client()
    if client is None:
        return jsonify({"ok": False, "error": "ANTHROPIC_API_KEY is not configured on the server. Add it to a .env file next to server.py."}), 503

    data = request.json or {}
    code = data.get('code', '') or ''
    project_name = data.get('project_name') or 'Project'
    language = data.get('language') or 'source'
    category_results = data.get('category_results') or {}

    if not code.strip():
        return jsonify({"ok": False, "error": "No source code provided"}), 400

    truncated = len(code) > AI_MAX_CODE_CHARS
    if truncated:
        code = code[:AI_MAX_CODE_CHARS]

    coverage_summary = "\n".join(
        f"- {cat}: {d.get('implemented', 0)}/{d.get('total', 0)} controls detected by pattern-matching ({d.get('pct', 0)}%)"
        for cat, d in category_results.items()
    ) or "(no pattern-matching results available)"

    user_content = (
        f"Project: {project_name}\n"
        f"Language: {language}\n\n"
        f"A separate static pattern-matching pass already scored ASVS 5.0 category coverage:\n{coverage_summary}\n\n"
        "Review the source code below and give your own independent, real assessment — base it on what you "
        "actually see in the code, not just the coverage percentages above.\n\n"
        + (f"[Note: source was truncated to {AI_MAX_CODE_CHARS} characters for this review]\n\n" if truncated else "")
        + f"```{language}\n{code}\n```"
    )

    try:
        response = client.messages.create(
            model=ANTHROPIC_MODEL,
            max_tokens=4000,
            output_config={"effort": "low", "format": {"type": "json_schema", "schema": AI_INSIGHTS_SCHEMA}},
            system=(
                "You are a security reviewer explaining findings to a non-expert reader. Write for someone "
                "with no security background: short sentences, everyday words, no jargon, no CWE numbers, "
                "no code snippets or function-name citations. Be direct and skimmable, not exhaustive."
            ),
            messages=[{"role": "user", "content": user_content}],
        )
    except anthropic.APIStatusError as e:
        return jsonify({"ok": False, "error": f"Claude API error: {e.message}"}), 502
    except anthropic.APIConnectionError:
        return jsonify({"ok": False, "error": "Could not reach the Claude API"}), 502

    if response.stop_reason == "refusal":
        return jsonify({"ok": False, "error": "The assessment request was declined by the model"}), 502

    text = next((b.text for b in response.content if b.type == "text"), None)
    if not text:
        return jsonify({"ok": False, "error": "Model returned no analysis"}), 502

    try:
        insights = json.loads(text)
    except json.JSONDecodeError:
        return jsonify({"ok": False, "error": "Model returned a malformed response"}), 502

    insights["ok"] = True
    return jsonify(insights)

DB = os.path.expanduser("~/asvs-compliance-maturity-analyzer/asvs.db")

def get_db():
    conn = sqlite3.connect(DB)
    conn.row_factory = sqlite3.Row
    return conn

def init_db():
    conn = get_db()
    conn.execute('''CREATE TABLE IF NOT EXISTS users (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        username TEXT UNIQUE NOT NULL,
        password TEXT NOT NULL,
        created_at TEXT DEFAULT CURRENT_TIMESTAMP
    )''')
    conn.execute('''CREATE TABLE IF NOT EXISTS scans (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        user_id INTEGER NOT NULL,
        project_name TEXT,
        files_count INTEGER,
        total_findings INTEGER,
        total_reqs INTEGER,
        overall_pct INTEGER,
        asvs_level TEXT,
        results_json TEXT,
        created_at TEXT DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY(user_id) REFERENCES users(id)
    )''')
    conn.commit()
    conn.close()

def hash_pw(pw):
    return hashlib.sha256(pw.encode()).hexdigest()

@app.route('/api/register', methods=['POST'])
def register():
    data = request.json
    try:
        conn = get_db()
        conn.execute("INSERT INTO users (username, password) VALUES (?, ?)",
                     (data['username'], hash_pw(data['password'])))
        conn.commit()
        user = conn.execute("SELECT id, username FROM users WHERE username=?", (data['username'],)).fetchone()
        conn.close()
        return jsonify({"ok": True, "user": {"id": user["id"], "username": user["username"]}})
    except sqlite3.IntegrityError:
        return jsonify({"ok": False, "error": "Username already exists"}), 400

@app.route('/api/login', methods=['POST'])
def login():
    data = request.json
    conn = get_db()
    user = conn.execute("SELECT id, username FROM users WHERE username=? AND password=?",
                        (data['username'], hash_pw(data['password']))).fetchone()
    conn.close()
    if user:
        return jsonify({"ok": True, "user": {"id": user["id"], "username": user["username"]}})
    return jsonify({"ok": False, "error": "Invalid credentials"}), 401

@app.route('/api/scans', methods=['POST'])
def save_scan():
    data = request.json
    conn = get_db()
    conn.execute("""INSERT INTO scans
        (user_id, project_name, files_count, total_findings, total_reqs, overall_pct, asvs_level, results_json)
        VALUES (?,?,?,?,?,?,?,?)""",
        (data['user_id'], data['project_name'], data['files_count'],
         data['total_findings'], data['total_reqs'], data['overall_pct'],
         data['asvs_level'], json.dumps(data['results'])))
    conn.commit()
    conn.close()
    return jsonify({"ok": True})

@app.route('/api/scans/<int:user_id>', methods=['GET'])
def get_scans(user_id):
    conn = get_db()
    scans = conn.execute("""SELECT id, project_name, files_count, total_findings,
        total_reqs, overall_pct, asvs_level, created_at
        FROM scans WHERE user_id=? ORDER BY created_at DESC LIMIT 20""",
        (user_id,)).fetchall()
    conn.close()
    return jsonify([dict(s) for s in scans])

@app.route('/api/scans/detail/<int:scan_id>', methods=['GET'])
def get_scan_detail(scan_id):
    conn = get_db()
    scan = conn.execute("SELECT * FROM scans WHERE id=?", (scan_id,)).fetchone()
    conn.close()
    if scan:
        d = dict(scan)
        d['results'] = json.loads(d['results_json'])
        return jsonify(d)
    return jsonify({"error": "Not found"}), 404

@app.route('/api/export/excel', methods=['POST'])
def export_excel():
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill as PF, Font as FN, Alignment as AL
    import os

    data = request.json
    results = data.get('categoryResults', {})

    TEMPLATE = '/home/kali/asvs-compliance-maturity-analyzer/ASVS-checklist-5.0.0.xlsx'
    if not os.path.exists(TEMPLATE):
        return jsonify({"error": "Template not found"}), 500

    wb = load_workbook(TEMPLATE)

    pass_fill = PF("solid", fgColor="C6EFCE")
    pass_font = FN(color="276221", bold=True, name="Calibri", size=10)
    fail_fill = PF("solid", fgColor="FFC7CE")
    fail_font = FN(color="9C0006", bold=True, name="Calibri", size=10)
    ctr_align = AL(horizontal="center", vertical="center")
    lft_align = AL(horizontal="left", vertical="center", wrap_text=True)

    remediation_map = {
        "2.1.1":"Enforce minimum 12 character password length on registration",
        "2.1.2":"Allow passwords up to 64+ characters without truncation",
        "2.1.3":"Do not truncate passwords during storage or comparison",
        "2.1.4":"Allow all printable Unicode characters in passwords",
        "2.1.5":"Provide self-service password change feature for users",
        "2.1.6":"Require current password before allowing password change",
        "2.1.7":"Integrate HaveIBeenPwned API to check against breached passwords",
        "2.1.8":"Add zxcvbn or similar password strength meter to registration UI",
        "2.1.9":"Remove all character composition rules — allow any character type",
        "2.1.10":"Remove periodic password rotation and password history requirements",
        "2.1.11":"Allow browser password managers and paste functionality",
        "2.1.12":"Add show/hide password toggle on login and registration forms",
        "2.2.1":"Add Flask-Limiter or express-rate-limit for brute force protection",
        "2.2.2":"Remove SMS-based auth or restrict to secondary factor only",
        "2.2.3":"Send security notifications for password changes and login attempts",
        "2.2.4":"Verify authenticator type is appropriate for risk level",
        "2.2.5":"Add replay-resistant MFA such as TOTP or FIDO2",
        "2.2.6":"Use time-based OTP instead of SMS for MFA",
        "2.2.7":"Add physical hardware token support for high-security accounts",
        "2.3.1":"Use secrets.token_hex for initial passwords with short expiry",
        "2.3.2":"Allow users to change initial passwords on first login",
        "2.3.3":"Use time-based password expiry for forgotten password flows",
        "2.4.1":"Use bcrypt argon2 or scrypt for password storage never MD5 or SHA1",
        "2.4.2":"Ensure salt is at least 32 bits and unique per credential",
        "2.4.3":"Set PBKDF2 iteration count to minimum 100000",
        "2.4.4":"Set bcrypt rounds to minimum 10 e.g. bcrypt.gensalt rounds=12",
        "2.4.5":"Use argon2id with minimum configuration per OWASP recommendation",
        "2.5.1":"Send recovery tokens via secure channel never in plain text email",
        "2.5.2":"Remove all secret question and hint-based recovery flows",
        "2.5.3":"Never use security questions as the only recovery mechanism",
        "2.5.4":"Audit for default credentials admin/admin root/root test/test",
        "2.5.5":"Invalidate all active sessions when password is changed or reset",
        "2.5.6":"Use secure random tokens for forgot password expire within 10 minutes",
        "2.5.7":"Require re-authentication before allowing MFA changes",
        "2.7.1":"Replace SMS OTP with TOTP speakeasy/pyotp or push notifications",
        "2.7.2":"Expire OOB authentication requests after 10 minutes",
        "2.7.3":"Allow only one active OOB authentication request at a time",
        "2.10.1":"Replace static API keys with short-lived tokens or OAuth2",
        "2.10.2":"Verify service accounts do not use default credentials",
        "2.10.3":"Store service credentials in secrets vault with access logging",
        "2.10.4":"Move all secrets to .env or vault never hardcode in source code",
        "3.1.1":"Never expose session tokens in URL parameters or error messages",
        "3.2.1":"Call session.regenerate() after successful login",
        "3.2.2":"Use cryptographically secure session ID with min 128 bits entropy",
        "3.2.3":"Store session tokens only in secure cookie never in localStorage",
        "3.2.4":"Use approved CSPRNG for session token generation",
        "3.3.1":"Destroy session server-side on logout not just clear client cookie",
        "3.3.2":"Add sliding session expiry and absolute timeout e.g. 8 hours",
        "3.3.3":"Invalidate all sessions when password is changed",
        "3.3.4":"Provide users ability to view and terminate all active sessions",
        "3.4.1":"Set Secure flag on all cookies Set-Cookie: id=x; Secure",
        "3.4.2":"Set HttpOnly flag on all cookies Set-Cookie: id=x; HttpOnly",
        "3.4.3":"Set SameSite=Strict or Lax on session cookies",
        "3.4.4":"Set cookie Path attribute to most specific path needed",
        "3.4.5":"Set cookie Domain attribute only when explicitly needed",
        "3.5.1":"Use stateless OAuth2 tokens instead of static API secrets",
        "3.5.2":"Use stateless JWT tokens avoid server-side session storage",
        "3.5.3":"Sign JWTs with RS256 or ES256 never use none algorithm",
        "3.7.1":"Re-authenticate users before sensitive actions like password change",
        "4.1.1":"Enforce access control server-side never trust client-side checks",
        "4.1.2":"Sanitize all user-supplied attributes used in access decisions",
        "4.1.3":"Apply least privilege users should only access their own resources",
        "4.1.5":"Return safe error on access control failure never expose internals",
        "4.2.1":"Validate resource ownership on every request to prevent IDOR",
        "4.2.2":"Add CSRF protection middleware to all state-changing routes",
        "4.3.1":"Protect admin routes with MFA requirement",
        "4.3.2":"Disable directory listing in web server config Options -Indexes",
        "4.3.3":"Add step-up authentication for high-value operations",
        "5.1.1":"Validate and reject duplicate or conflicting HTTP parameters",
        "5.1.2":"Use allowlist for mass assignment never bind all request params",
        "5.1.3":"Use allowlist validation reject anything not explicitly permitted",
        "5.1.4":"Validate JSON/XML against a strict schema using joi or zod",
        "5.1.5":"Maintain URL redirect allowlist to block open redirects",
        "5.2.1":"Use DOMPurify or bleach to sanitize all HTML input",
        "5.2.2":"Sanitize unstructured data for allowed characters and length",
        "5.2.3":"Sanitize user input before passing to mail systems",
        "5.2.4":"Remove all eval() calls use JSON.parse() and safe alternatives",
        "5.2.5":"Sanitize or sandbox user-supplied template content",
        "5.2.6":"Validate outbound URLs against allowlist to prevent SSRF",
        "5.2.7":"Sanitize SVG content to prevent XSS via inline scripts",
        "5.2.8":"Sanitize Markdown or BBCode before rendering to prevent XSS",
        "5.3.1":"Use context-aware output encoding for HTML JS and URL contexts",
        "5.3.2":"Preserve character encoding use UTF-8 consistently",
        "5.3.3":"Use context-aware output escaping to prevent XSS",
        "5.3.4":"Use parameterized queries or ORM for all database operations",
        "5.3.5":"Use SQL escaping when parameterized queries are unavailable",
        "5.3.6":"Protect against JSON injection and JavaScript expression evaluation",
        "5.3.7":"Use LDAP escaping to prevent LDAP injection",
        "5.3.8":"Use subprocess with list args never shell=True with user input",
        "5.3.9":"Protect against Local File Inclusion LFI attacks",
        "5.3.10":"Use XML/XPath libraries with entity expansion disabled",
        "5.5.1":"Add HMAC signature to serialized objects to detect tampering",
        "5.5.2":"Use yaml.safe_load() and defusedxml disable entity expansion",
        "5.5.3":"Use JSON.parse() exclusively avoid pickle for untrusted data",
        "5.5.4":"Use JSON.parse() instead of eval() for JSON parsing",
        "6.1.1":"Encrypt PII fields at rest using AES-256-GCM before storing",
        "6.1.2":"Encrypt health records using approved algorithms AES-256",
        "6.1.3":"Encrypt financial data fields at rest with column-level encryption",
        "6.2.1":"Catch all crypto exceptions never expose padding oracle errors",
        "6.2.2":"Use AES-256-GCM ChaCha20-Poly1305 or RSA-2048+ only",
        "6.2.3":"Configure IV/nonce correctly use GCM or CBC with random IV",
        "6.2.4":"Make crypto algorithms configurable for easy rotation",
        "6.2.5":"Replace MD5/SHA1 with SHA-256 or bcrypt/argon2 for passwords",
        "6.2.6":"Generate a fresh IV/nonce for every encryption operation",
        "6.2.7":"Use authenticated encryption GCM/CCM to prevent tampering",
        "6.2.8":"Use constant-time comparison for cryptographic operations",
        "6.3.1":"Use os.urandom() or secrets module never random.random()",
        "6.3.2":"Use uuid.uuid4() for all GUID generation",
        "6.3.3":"Ensure sufficient entropy under high load conditions",
        "6.4.1":"Store secrets in HashiCorp Vault AWS KMS or Azure Key Vault",
        "6.4.2":"Never load raw key material into app memory use vault API calls",
        "7.1.1":"Strip passwords and tokens from logs using log filtering middleware",
        "7.1.2":"Exclude PII from logs mask email and phone before logging",
        "7.1.3":"Log all auth events login logout failed attempts with timestamp",
        "7.1.4":"Include user ID IP timestamp and action in every security log entry",
        "7.2.1":"Log all authentication decisions never log passwords or tokens",
        "7.2.2":"Log all access control denials with user resource and action",
        "7.3.1":"Encode log output to prevent CRLF/newline log injection",
        "7.3.2":"Protect log viewer from log injection attacks",
        "7.3.3":"Store logs in append-only storage restrict write access",
        "7.4.1":"Show generic error messages to users log details server-side only",
        "7.4.2":"Wrap all external calls in try/except handle all exception types",
        "7.4.3":"Add global exception handler Flask @app.errorhandler(Exception)",
        "8.1.1":"Prevent sensitive data caching in server-side load balancers",
        "8.1.2":"Purge cached sensitive data after authorized user accesses it",
        "8.1.3":"Minimize parameters in requests remove unnecessary hidden fields",
        "8.1.4":"Implement request rate limiting to prevent DoS attacks",
        "8.2.1":"Set Cache-Control no-store no-cache on all sensitive responses",
        "8.2.2":"Never store sensitive data in localStorage or sessionStorage",
        "8.2.3":"Clear all sensitive DOM data and storage on session logout",
        "8.3.1":"Pass sensitive data in request body or headers never in URL params",
        "8.3.2":"Provide users a method to export or delete their personal data",
        "8.3.3":"Obtain user consent before collecting personal information",
        "8.3.4":"Document and classify all sensitive data handled by the application",
        "8.3.5":"Log access to sensitive data for audit purposes",
        "8.3.6":"Overwrite sensitive memory with zeros when no longer needed",
        "8.3.7":"Encrypt sensitive fields using AES-256-GCM before storage",
        "8.3.8":"Implement data retention policy auto-delete outdated sensitive data",
        "9.1.1":"Enforce HTTPS redirect all HTTP to HTTPS with 301",
        "9.1.2":"Disable weak cipher suites use only TLS_AES_256_GCM_SHA384",
        "9.1.3":"Set minimum TLS 1.2 disable TLS 1.0 and 1.1 in server config",
        "9.2.1":"Use certificates from trusted CA validate cert chain",
        "9.2.2":"Use TLS for all internal service-to-service communication",
        "9.2.3":"Set verify=True on all requests never disable cert verification",
        "9.2.4":"Enable OCSP stapling in nginx/Apache web server configuration",
        "9.2.5":"Log all backend TLS connection failures for monitoring",
        "12.1.1":"Set MAX_CONTENT_LENGTH to limit file upload size",
        "12.1.2":"Check compressed files for zip bomb before extraction",
        "12.1.3":"Enforce per-user file count and total size quotas",
        "12.2.1":"Validate file content type using magic bytes not just extension",
        "12.2.2":"Validate structured file format before processing PDF DOCX",
        "12.2.3":"Validate XML files against schema before processing",
        "12.3.1":"Use os.path.basename() to prevent path traversal in filenames",
        "12.3.2":"Validate filenames against allowlist to prevent LFI",
        "12.3.3":"Validate file URLs against allowlist to prevent SSRF via RFI",
        "12.3.4":"Set Content-Disposition header to prevent RFD attacks",
        "12.3.5":"Never pass user filenames directly to OS or system calls",
        "12.4.1":"Store uploads outside web root in non-executable directory",
        "12.4.2":"Scan uploaded files with antivirus before processing",
        "12.5.1":"Restrict web server to serve only allowed file extensions",
        "12.5.2":"Never execute uploaded files as HTML or JavaScript",
        "12.6.1":"Configure SSRF allowlist for server-side URL fetch operations",
        "13.1.1":"Use consistent parsers across all application components",
        "13.1.2":"Validate all API input using OpenAPI/Swagger schema",
        "13.1.3":"Remove API keys from URLs use Authorization header instead",
        "13.1.4":"Enforce authorization at both URI and resource level",
        "13.1.5":"Reject requests with unexpected Content-Type headers",
        "13.2.1":"Restrict HTTP methods block unused verbs DELETE PUT",
        "13.2.2":"Validate all JSON request bodies against JSON schema",
        "13.2.3":"Add CSRF token validation to all state-changing API endpoints",
        "13.2.4":"Validate inbound JSON with allowlist of expected fields",
        "13.2.5":"Check Content-Type header on all API endpoints",
        "13.2.6":"Verify message headers and payload integrity with TLS",
        "13.3.1":"Validate XML with XSD schema before processing for XXE prevention",
        "13.3.2":"Sign SOAP messages with WS-Security for transport integrity",
        "13.4.1":"Implement GraphQL query depth and complexity limits",
        "13.4.2":"Implement GraphQL authorization at business logic layer",
        "14.1.1":"Use separate build test and production environments",
        "14.1.2":"Use Infrastructure-as-Code for all environment configuration",
        "14.1.3":"Deploy using CI/CD pipeline with security scanning stages",
        "14.1.4":"Restrict production deployment access to authorized personnel",
        "14.1.5":"Record and alert on all production deployment events",
        "14.2.1":"Use dependency checker OWASP Dependency-Check or Snyk in CI/CD",
        "14.2.2":"Remove all unused dependencies features and sample code",
        "14.2.3":"Use Subresource Integrity SRI for all CDN-hosted assets",
        "14.2.4":"Use only packages from trusted maintained repositories",
        "14.2.5":"Verify third-party library integrity with checksums",
        "14.2.6":"Reduce attack surface by removing unneeded features and modules",
        "14.3.1":"Configure custom error pages never expose stack traces",
        "14.3.2":"Set DEBUG=False and NODE_ENV=production in all production configs",
        "14.3.3":"Remove server version headers X-Powered-By and Server",
        "14.4.1":"Set Content-Type header on all HTTP responses",
        "14.4.2":"Set Content-Disposition attachment on all API JSON responses",
        "14.4.3":"Add Content-Security-Policy default-src self header",
        "14.4.4":"Add X-Content-Type-Options nosniff to all responses",
        "14.4.5":"Add HSTS header Strict-Transport-Security max-age=31536000",
        "14.4.6":"Add Referrer-Policy no-referrer or strict-origin header",
        "14.4.7":"Add X-Frame-Options DENY or CSP frame-ancestors none",
        "14.5.1":"Accept only HTTP methods in use reject others with 405",
        "14.5.2":"Never use Origin header for authentication or access control",
        "14.5.3":"Set CORS Access-Control-Allow-Origin to specific domain never wildcard",
        "14.5.4":"Authenticate proxy-added headers bearer tokens SSO headers",
    }

    sheet_map = {
        "Encoding and Sanitization": "Encoding and Sanitization",
        "Validation and Business Logic": "Validation and Business Logic",
        "Web Frontend Security": "Web Frontend Security",
        "API and Web Service": "API and Web Service",
        "File Handling": "File Handling",
        "Authentication": "Authentication",
        "Session Management": "Session Management",
        "Authorization": "Authorization",
        "Self-contained Tokens": "Self-contained Tokens",
        "OAuth and OIDC": "OAuth and OIDC",
        "Cryptography": "Cryptography",
        "Secure Communication": "Secure Communication",
        "Configuration": "Configuration",
        "Data Protection": "Data Protection",
        "Secure Coding and Architecture": "Secure Coding and Architecture",
        "Security Logging and Error Handling": "Security Logging & Errors",
        "WebRTC": "WebRTC",
    }

    # Remove unwanted sheets

    for cat, cat_data in results.items():
        sheet_name = sheet_map.get(cat)
        if not sheet_name or sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]

        # Hide NIST (col E) and Tool Used (col J)
        ws.column_dimensions["E"].hidden = True
        ws.column_dimensions["J"].hidden = True
        ws.column_dimensions["J"].width = 0
        # Set column widths
        ws.column_dimensions["A"].width = 18
        ws.column_dimensions["B"].width = 10
        ws.column_dimensions["C"].width = 8
        ws.column_dimensions["D"].width = 8
        ws.column_dimensions["F"].width = 55
        ws.column_dimensions["G"].width = 20
        ws.column_dimensions["H"].width = 45
        ws.column_dimensions["I"].width = 40

        req_map = {str(r.get("id","")).strip(): r for r in cat_data.get("reqs",[])}

        for row in ws.iter_rows(min_row=2):
            ws.row_dimensions[row[0].row].height = 60
            # Apply wrap text to ALL cells in row
            for cell in row:
                if cell.alignment:
                    cell.alignment = AL(wrap_text=True, vertical="top", horizontal="left")
                else:
                    cell.alignment = AL(wrap_text=True, vertical="top", horizontal="left")
            req_id = str(row[1].value).strip() if row[1].value else ""
            if not req_id or req_id in ["None","nan",""]:
                continue
            req = req_map.get(req_id)
            implemented = req.get("implemented", False) if req else False
            finding = req.get("finding", {}) if req else {}

            valid_cell   = row[6]
            src_cell     = row[7]
            comment_cell = row[8]
            level_ref = f"ASVS Level {req.get('level','')}" if req and req.get("level") else ""

            wrong_impl = req.get("wrongImplementation", False) if req else False
            # rmap for wrong implementation guidance
            rmap = {
                "2.4.1": ("Use bcrypt/argon2 for password storage", "Replace with bcrypt.hashpw(pwd.encode(), bcrypt.gensalt(12))"),
                "2.4.4": ("Set bcrypt rounds >= 10", "Change cost factor to bcrypt.gensalt(rounds=12)"),
                "3.4.1": ("Set Secure flag on cookies", "Add secure=True to cookie config"),
                "3.4.2": ("Set HttpOnly flag on cookies", "Add httponly=True to cookie config"),
                "3.4.3": ("Set SameSite=Strict on cookies", "Add samesite=Strict to cookie config"),
                "5.3.4": ("Use parameterized queries", "Replace string SQL with SQLAlchemy text() and named params"),
                "6.2.2": ("Use AES-256-GCM", "Replace weak cipher with AES-256-GCM"),
                "6.3.1": ("Use secrets module", "Replace Math.random/rand with secrets.token_hex()"),
                "9.2.3": ("Enable SSL verification", "Set verify=True in all requests calls"),
                "14.3.2": ("Disable debug mode", "Set DEBUG=False in production config"),
            }
            if implemented and not wrong_impl:
                valid_cell.value = "Valid"
                valid_cell.fill  = pass_fill
                valid_cell.font  = pass_font
                valid_cell.alignment = ctr_align
                # Source Code Reference - real line number and code
                line_num = finding.get("lineNumber") if finding else None
                line_content = finding.get("lineContent") if finding else None
                note = finding.get("note","Control detected") if finding else "Control detected"
                conf = finding.get("confidence","").upper() if finding else ""
                if line_num and line_content:
                    src_cell.value = f"Line {line_num}: {line_content}"
                    comment_cell.value = f"[{conf}] {note} detected at line {line_num}"
                else:
                    src_cell.value = note
                    comment_cell.value = f"Confidence: {conf} — {note}" if conf else "Pattern match confirmed"
                src_cell.alignment = lft_align
            elif wrong_impl:
                # CASE 2: Wrong implementation
                valid_cell.value = "Wrong Implementation"
                valid_cell.fill  = PF("solid", fgColor="FFEB9C")
                valid_cell.font  = FN(color="9C6500", bold=True, name="Calibri", size=10)
                valid_cell.alignment = ctr_align
                line_num = finding.get("lineNumber") if finding else None
                line_content = finding.get("lineContent") if finding else None
                wrong_note = finding.get("note","Wrong implementation") if finding else ""
                guidance, fix = rmap.get(req_id, ("Review per ASVS 5.0","Fix required"))
                if line_num and line_content:
                    src_cell.value = f"Line {line_num}: {line_content}"
                    comment_cell.value = f"WRONG at line {line_num} — {wrong_note}. Fix: {fix[:80]} | {level_ref}"
                else:
                    src_cell.value = guidance
                    comment_cell.value = f"Wrong implementation — {wrong_note} | {level_ref}"
                src_cell.alignment = lft_align
            else:
                valid_cell.value = "Not Valid"
                valid_cell.fill  = fail_fill
                valid_cell.font  = fail_font
                valid_cell.alignment = ctr_align
                src_cell.value = remediation_map.get(req_id, "Implement: " + (req.get("requirement", "Review ASVS 5.0.0 spec for this requirement") if req else "Review ASVS 5.0.0 spec for this requirement"))
                src_cell.alignment = lft_align
                unique_comments = {
                    "2.1.1":"Add min length validation to registration handler",
                    "2.1.2":"Remove upper limit on password field length",
                    "2.1.7":"Call pwnedpasswords API on registration and password change",
                    "2.1.8":"Add zxcvbn.js password strength meter to registration UI",
                    "2.2.1":"Install Flask-Limiter on login endpoint — 5 attempts per minute",
                    "2.2.2":"Replace SMS OTP with TOTP authenticator app",
                    "2.3.1":"Use secrets.token_hex(8) for temp passwords with 24h expiry",
                    "2.4.1":"Replace MD5/SHA1 — use bcrypt.hashpw(pwd.encode(), bcrypt.gensalt(12))",
                    "2.4.2":"bcrypt generates salt automatically — verify not using custom salt",
                    "2.4.3":"Set PBKDF2 iterations to 100000 minimum",
                    "2.4.4":"Set bcrypt cost factor to minimum 10 rounds",
                    "2.5.1":"Send reset link via HTTPS — never include token in email body",
                    "2.5.2":"Delete all security question fields from registration form",
                    "2.5.4":"Scan all config files — remove admin/admin root/root defaults",
                    "2.5.6":"Use secrets.token_urlsafe(32) — expire reset tokens in 10 min",
                    "2.7.1":"Replace SMS — integrate pyotp TOTP authenticator app",
                    "2.10.1":"Implement JWT bearer tokens for service authentication",
                    "2.10.4":"Use python-dotenv — move all secrets to .env file",
                    "3.1.1":"Never append session ID to URL — use Set-Cookie header only",
                    "3.2.1":"Call session.clear() then assign new session after login",
                    "3.2.2":"Use secrets.token_hex(32) for 256-bit session ID entropy",
                    "3.3.1":"Call session.clear() and delete_cookie() on logout endpoint",
                    "3.3.2":"Set PERMANENT_SESSION_LIFETIME = timedelta(hours=8)",
                    "3.4.1":"Add secure=True to all session cookie configuration",
                    "3.4.2":"Add httponly=True to all session cookie configuration",
                    "3.4.3":"Add samesite=Strict to all session cookie configuration",
                    "3.4.4":"Set cookie path to most specific path — not root /",
                    "3.5.2":"Use jwt.encode with RS256 algorithm for stateless tokens",
                    "3.5.3":"Set algorithm=RS256 — never use alg=none in JWT signing",
                    "3.7.1":"Require password confirmation before sensitive operations",
                    "4.1.1":"Add @login_required decorator to all protected routes",
                    "4.1.2":"Validate user role from JWT before access control decisions",
                    "4.1.3":"Add resource.user_id == current_user.id check on all resources",
                    "4.2.1":"Check db.get(id).owner == user.id before serving resource",
                    "4.2.2":"Install Flask-WTF and enable CSRF protection globally",
                    "4.3.1":"Add @require_mfa decorator to all admin routes",
                    "4.3.2":"Add autoindex off to nginx or Options -Indexes to Apache",
                    "5.1.3":"Use Marshmallow schema with strict field definitions",
                    "5.1.4":"Add jsonschema.validate() to all API request handlers",
                    "5.1.5":"Validate redirect_url against ALLOWED_URLS allowlist",
                    "5.2.1":"Use bleach.clean(input, tags=ALLOWED_TAGS) on all HTML",
                    "5.2.4":"Replace eval() with json.loads() throughout codebase",
                    "5.2.6":"Validate outbound URLs against allowed domain allowlist",
                    "5.3.3":"Enable Jinja2 autoescaping — render_template with autoescape",
                    "5.3.4":"Use SQLAlchemy text() with named parameters for all queries",
                    "5.3.8":"Replace os.system() with subprocess.run(list, shell=False)",
                    "5.5.3":"Replace pickle.loads() with json.loads() for user data",
                    "6.1.1":"Encrypt PII: Fernet(key).encrypt(value.encode()) before store",
                    "6.2.1":"Wrap all crypto in try/except — return generic error message",
                    "6.2.2":"Use AES-256-GCM: Cipher(algorithms.AES(key), modes.GCM(iv))",
                    "6.2.5":"Replace hashlib.md5/sha1 with hashlib.sha256 or bcrypt",
                    "6.2.6":"Generate new IV each time: iv = os.urandom(12) — never reuse",
                    "6.3.1":"Replace random.random() with secrets.token_hex(32) for tokens",
                    "6.3.2":"Replace uuid.uuid1() with uuid.uuid4() for all GUIDs",
                    "6.4.1":"Move keys to HashiCorp Vault or AWS KMS — never in source",
                    "6.4.2":"Use Vault transit API — never load raw key into app memory",
                    "7.1.1":"Add log filter to redact Authorization and password fields",
                    "7.1.2":"Mask PII: email[:3]+'***@'+domain before writing to log",
                    "7.1.3":"Log: logger.info('LOGIN user=%s ip=%s result=%s', u, ip, r)",
                    "7.2.1":"Log auth decisions — filter out password and token values",
                    "7.3.1":"Escape newlines: log_msg.replace(chr(10), '\\n') before log",
                    "7.4.1":"Return generic 500 to user — log full traceback server-side",
                    "7.4.2":"Wrap DB/API calls in try/except Exception as e: log(e)",
                    "7.4.3":"Add @app.errorhandler(Exception) global exception handler",
                    "8.2.1":"Set Cache-Control: no-store on all sensitive API responses",
                    "8.2.2":"Move sensitive data from localStorage to httpOnly cookies",
                    "8.3.1":"Move API keys from URL params to Authorization header",
                    "8.3.7":"Encrypt PII columns with AES-256-GCM before db.save()",
                    "9.1.1":"Add before_request redirect HTTP to HTTPS with 301",
                    "9.1.2":"Set nginx ssl_ciphers ECDHE-RSA-AES256-GCM-SHA384 only",
                    "9.1.3":"Set context.minimum_version = ssl.TLSVersion.TLSv1_2",
                    "9.2.1":"Set verify=True and ca_bundle in all requests calls",
                    "9.2.3":"Remove verify=False from all requests.get/post calls",
                    "9.2.4":"Add ssl_stapling on; ssl_stapling_verify on; to nginx",
                    "12.1.1":"Set app.config[MAX_CONTENT_LENGTH] = 10 * 1024 * 1024",
                    "12.1.2":"Check decompressed size — reject if > 100x compressed",
                    "12.3.1":"Apply os.path.basename() to all user-supplied filenames",
                    "12.4.1":"Set UPLOAD_FOLDER outside web root — /var/uploads",
                    "13.1.3":"Move ?api_key=X to Authorization: Bearer X header",
                    "13.2.2":"Add jsonschema.validate(request.json, schema) to endpoints",
                    "13.2.3":"Enable Flask-WTF CSRF or validate X-CSRF-Token header",
                    "13.2.5":"Return 415 if Content-Type is not application/json",
                    "14.2.1":"Add safety check to CI/CD: pip install safety && safety check",
                    "14.2.3":"Add integrity=sha384-X crossorigin=anonymous to CDN scripts",
                    "14.3.1":"Add @app.errorhandler(404) and @app.errorhandler(500) pages",
                    "14.3.2":"Set FLASK_DEBUG=0 and FLASK_ENV=production in .env",
                    "14.3.3":"Add server_tokens off to nginx — remove X-Powered-By header",
                    "14.4.1":"Set explicit Content-Type on all Flask response objects",
                    "14.4.3":"Add response.headers[CSP] = default-src self in after_request",
                    "14.4.4":"Add response.headers[X-Content-Type-Options] = nosniff",
                    "14.4.5":"Add Strict-Transport-Security: max-age=31536000 header",
                    "14.4.6":"Add response.headers[Referrer-Policy] = no-referrer",
                    "14.4.7":"Add response.headers[X-Frame-Options] = DENY",
                    "14.5.1":"Specify methods=[GET,POST] in all route decorators",
                    "14.5.3":"Set CORS origins=[https://yourdomain.com] — never wildcard",
                }
                cwe_text = f" | {level_ref}" if level_ref else ""
                comment_cell.value = unique_comments.get(req_id, "Review ASVS 5.0 spec for " + req_id + " implementation guidance") + cwe_text

    # Update ASVS Results sheet
    if "ASVS Results" in wb.sheetnames:
        ws_r = wb["ASVS Results"]
        from openpyxl.styles import PatternFill as PF2, Font as FN2, Alignment as AL2
        lv2_f = PF2("solid",fgColor="C6EFCE"); lv2_ft = FN2(color="276221",bold=True,name="Calibri",size=10)
        lv1_f = PF2("solid",fgColor="FFEB9C"); lv1_ft = FN2(color="9C6500",bold=True,name="Calibri",size=10)
        bl1_f = PF2("solid",fgColor="FFC7CE"); bl1_ft = FN2(color="9C0006",bold=True,name="Calibri",size=10)
        na_f  = PF2("solid",fgColor="F2F2F2"); na_ft  = FN2(color="595959",italic=True,name="Calibri",size=10)
        ctr2  = AL2(horizontal="center",vertical="center")
        # Use our mapped totals (103) not full ASVS sheet totals (283)
        t_found = sum(cd.get("implemented",0) for cd in results.values())
        t_reqs = sum(cd.get("total",0) for cd in results.values())
        t_all_reqs = [r for cd in results.values() for r in cd.get("reqs", [])]
        t_pct, t_level = compute_weighted_score(t_all_reqs)

        # Override the Excel formula cells in Total row with our values
        for row in ws_r.iter_rows(min_row=2):
            cat_name = str(row[0].value).strip().lower() if row[0].value else ""
            if cat_name == "total":
                # Replace Excel SUM formulas with our mapped values
                row[1].value = t_found   # Valid criteria
                row[2].value = t_reqs    # Total criteria (103)
                row[3].value = round(t_pct, 2)  # Percentage
                break
        from openpyxl.styles import Alignment as AL2
        skip_cats = ['architecture', 'malicious', 'business']
        rows_to_delete = []
        for row in ws_r.iter_rows(min_row=2):
            cat_name = str(row[0].value).strip() if row[0].value else ""
            if not cat_name: continue
            # Delete architecture malicious business rows
            if any(s in cat_name.lower() for s in skip_cats):
                rows_to_delete.append(row[0].row)
                continue
            # Fix VALUE errors - replace formulas with plain values
            for cell in row:
                if cell.value and str(cell.value).startswith('='):
                    cell.value = None
                # Apply wrap text to all cells
                cell.alignment = AL2(wrap_text=True, vertical="top", horizontal="left")

        for row_num in sorted(rows_to_delete, reverse=True):
            ws_r.delete_rows(row_num)

        # Fix column widths in ASVS Results
        ws_r.column_dimensions["A"].width = 28
        ws_r.column_dimensions["B"].width = 14
        ws_r.column_dimensions["C"].width = 14
        ws_r.column_dimensions["D"].width = 16
        ws_r.column_dimensions["E"].width = 18

        for row in ws_r.iter_rows(min_row=2):
            cat_name = str(row[0].value).strip() if row[0].value else ""
            if not cat_name: continue
            if cat_name.lower() == "total":
                if t_level=="Level 2": lvl_cell.value="Level 2"; lvl_cell.fill=lv2_f; lvl_cell.font=lv2_ft
                elif t_level=="Level 1": lvl_cell.value="Level 1"; lvl_cell.fill=lv1_f; lvl_cell.font=lv1_ft
                else: lvl_cell.value="Below Level 1"; lvl_cell.fill=bl1_f; lvl_cell.font=bl1_ft
                lvl_cell.alignment=ctr2; continue
            lvl_cell = row[4]
            matched = None
            for our_cat in sheet_map:
                if cat_name.lower() in our_cat.lower() or our_cat.lower() in cat_name.lower():
                    matched = our_cat; break
            if matched and matched in results:
                cd = results[matched]
                imp = cd.get("implemented",0); tot = cd.get("total",0)
                pct, cat_level = compute_weighted_score(cd.get("reqs", []))
                # Override Excel formula cells with our mapped values
                row[1].value = imp   # Valid criteria (our mapped count)
                row[2].value = tot   # Total criteria (our mapped total)
                row[3].value = round(pct, 2)  # Our Level-weighted percentage
                if cat_level=="Level 2":
                    lvl_cell.value="Level 2"; lvl_cell.fill=lv2_f; lvl_cell.font=lv2_ft
                elif cat_level=="Level 1":
                    lvl_cell.value="Level 1"; lvl_cell.fill=lv1_f; lvl_cell.font=lv1_ft
                else:
                    lvl_cell.value="Below Level 1"; lvl_cell.fill=bl1_f; lvl_cell.font=bl1_ft
                lvl_cell.alignment=ctr2
            else:
                lvl_cell.value="Not Assessed"; lvl_cell.fill=na_f; lvl_cell.font=na_ft
                lvl_cell.alignment=ctr2

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="ASVS-Compliance-Maturity-Report.xlsx")



@app.route('/api/export/pdf', methods=['POST'])
def export_pdf():
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.styles import ParagraphStyle
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable, PageBreak
        from reportlab.lib.units import cm
        from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT, TA_JUSTIFY
        import datetime

        data = request.json
        results = data.get('categoryResults', {})
        project = data.get('project_name', 'ASVS Analysis')
        ai_insights = data.get('ai_insights')  # real Claude output from the UI's AI assessment, if it ran
        buf = io.BytesIO()

        doc = SimpleDocTemplate(buf, pagesize=A4,
            topMargin=2*cm, bottomMargin=2*cm,
            leftMargin=2*cm, rightMargin=2*cm,
            title="ASVS Compliance & Maturity Analyzer - " + project, author="ASVS Compliance & Maturity Analyzer")

        NAVY   = colors.HexColor("#1F3864")
        BLUE   = colors.HexColor("#2E75B6")
        LBLUE  = colors.HexColor("#D6E4F0")
        LLBLUE = colors.HexColor("#EEF4FF")
        GREEN  = colors.HexColor("#375623")
        LGREEN = colors.HexColor("#E2EFDA")
        RED    = colors.HexColor("#9C0006")
        LRED   = colors.HexColor("#FFC7CE")
        AMBER  = colors.HexColor("#9C6500")
        WHITE  = colors.white
        GRAY   = colors.HexColor("#595959")
        LGRAY  = colors.HexColor("#F2F2F2")
        DGRAY  = colors.HexColor("#333333")
        MGRAY  = colors.HexColor("#CCCCCC")

        def S(name, **kw):
            return ParagraphStyle(name, **kw)

        title_s  = S("ts", fontSize=22, fontName="Helvetica-Bold", textColor=WHITE, alignment=TA_CENTER, leading=28)
        sub_s    = S("ss", fontSize=11, fontName="Helvetica", textColor=colors.HexColor("#DCE6F1"), alignment=TA_CENTER)
        h1_s     = S("h1", fontSize=14, fontName="Helvetica-Bold", textColor=NAVY, spaceBefore=12, spaceAfter=5)
        h2_s     = S("h2", fontSize=11, fontName="Helvetica-Bold", textColor=BLUE, spaceBefore=8, spaceAfter=3)
        body_s   = S("bs", fontSize=10, fontName="Helvetica", textColor=DGRAY, spaceAfter=4, leading=15, alignment=TA_JUSTIFY)
        small_s  = S("sm", fontSize=8, fontName="Helvetica", textColor=GRAY, alignment=TA_CENTER)
        meta_k   = S("mk", fontSize=10, fontName="Helvetica-Bold", textColor=NAVY)
        meta_v   = S("mv", fontSize=10, fontName="Helvetica", textColor=DGRAY)
        footer_s = S("fs", fontSize=8, fontName="Helvetica-Oblique", textColor=GRAY, alignment=TA_CENTER, leading=10)

        story = []
        now = datetime.datetime.now().strftime("%d %B %Y  %H:%M")
        total_found = sum(cd.get("implemented",0) for cd in results.values())
        total_reqs  = sum(cd.get("total",0) for cd in results.values())
        all_reqs_flat = [r for cd in results.values() for r in cd.get("reqs", [])]
        overall, lvl = compute_weighted_score(all_reqs_flat)
        lvl_color   = GREEN if lvl=="Level 2" else AMBER if lvl=="Level 1" else RED
        cats_active = len([c for c in results if results[c].get("implemented",0)>0])

        # COVER BANNER
        banner = Table([
            [Paragraph("ASVS Compliance & Maturity Analyzer", title_s)],
            [Paragraph("Application Security Verification Standard Report", sub_s)],
            [Paragraph("OWASP ASVS 5.0  ·  Automated Security Assessment  ·  " + now, sub_s)]
        ], colWidths=[17*cm])
        banner.setStyle(TableStyle([
            ("BACKGROUND",(0,0),(-1,-1),NAVY),
            ("TOPPADDING",(0,0),(-1,-1),18),
            ("BOTTOMPADDING",(0,0),(-1,-1),18),
            ("LEFTPADDING",(0,0),(-1,-1),24),
        ]))
        story.append(banner)
        story.append(Spacer(1,0.4*cm))

        # META TABLE
        meta = Table([
            [Paragraph("Project", meta_k),         Paragraph(project, meta_v)],
            [Paragraph("Standard", meta_k),         Paragraph("OWASP Application Security Verification Standard 5.0", meta_v)],
            [Paragraph("Tool", meta_k),             Paragraph("ASVS Compliance & Maturity Analyzer", meta_v)],
            [Paragraph("Overall Coverage", meta_k), Paragraph(str(overall)+"%", S("ov",fontSize=10,fontName="Helvetica-Bold",textColor=lvl_color))],
            [Paragraph("Maturity Level", meta_k),   Paragraph(lvl, S("lv",fontSize=10,fontName="Helvetica-Bold",textColor=lvl_color))],
        ], colWidths=[4.5*cm,12.5*cm])
        meta.setStyle(TableStyle([
            ("BACKGROUND",(0,0),(0,-1),LBLUE),
            ("BACKGROUND",(1,0),(1,-1),WHITE),
            ("GRID",(0,0),(-1,-1),0.5,MGRAY),
            ("TOPPADDING",(0,0),(-1,-1),7),
            ("BOTTOMPADDING",(0,0),(-1,-1),7),
            ("LEFTPADDING",(0,0),(-1,-1),10),
            ("VALIGN",(0,0),(-1,-1),"MIDDLE"),
        ]))
        story.append(meta)
        story.append(Spacer(1,0.4*cm))
        story.append(HRFlowable(width="100%",thickness=2,color=NAVY))
        story.append(Spacer(1,0.3*cm))

        # EXECUTIVE SUMMARY
        story.append(Paragraph("Executive Summary", h1_s))
        scores = Table([
            [Paragraph(str(total_found), S("n",fontSize=26,fontName="Helvetica-Bold",textColor=BLUE,alignment=TA_CENTER)),
             Paragraph(str(total_reqs),  S("n",fontSize=26,fontName="Helvetica-Bold",textColor=NAVY,alignment=TA_CENTER)),
             Paragraph(str(overall)+"%", S("n",fontSize=26,fontName="Helvetica-Bold",textColor=lvl_color,alignment=TA_CENTER)),
             Paragraph(lvl,              S("n",fontSize=14,fontName="Helvetica-Bold",textColor=lvl_color,alignment=TA_CENTER)),
             Paragraph(str(cats_active)+"/"+str(len(results)), S("n",fontSize=26,fontName="Helvetica-Bold",textColor=BLUE,alignment=TA_CENTER))],
            [Paragraph("Controls Found",small_s),
             Paragraph("Total Requirements",small_s),
             Paragraph("Coverage",small_s),
             Paragraph("Maturity Level",small_s),
             Paragraph("Categories Active",small_s)],
        ], colWidths=[3.4*cm]*5)
        scores.setStyle(TableStyle([
            ("BACKGROUND",(0,0),(-1,-1),LLBLUE),
            ("TOPPADDING",(0,0),(-1,-1),12),
            ("BOTTOMPADDING",(0,0),(-1,-1),8),
            ("ALIGN",(0,0),(-1,-1),"CENTER"),
            ("VALIGN",(0,0),(-1,-1),"MIDDLE"),
            ("LINEABOVE",(0,0),(-1,0),0,WHITE),
            ("LINEBELOW",(0,-1),(-1,-1),0,WHITE),
            ("LINEBEFORE",(0,0),(0,-1),0,WHITE),
            ("LINEAFTER",(-1,0),(-1,-1),0,WHITE),
            ("INNERGRID",(0,0),(-1,-1),0,WHITE),
        ]))
        story.append(scores)
        story.append(Spacer(1,0.1*cm))

        top_cats  = [c for c,d in sorted(results.items(),key=lambda x:x[1].get("pct",0),reverse=True) if d.get("implemented",0)>0][:2]
        weak_cats = [c for c,d in sorted(results.items(),key=lambda x:x[1].get("pct",0)) if d.get("pct",0)<40][:2]
        summary = ("This automated ASVS 5.0 security assessment analyzed the provided source code and detected "
            + str(total_found) + " implemented security controls out of " + str(total_reqs) + " mapped requirements, "
            "achieving " + str(overall) + "% overall coverage. "
            + ("Strongest areas include " + " and ".join(top_cats) + ". " if top_cats else "")
            + ("Priority remediation: " + " and ".join(weak_cats) + " show lowest coverage. " if weak_cats else "")
            + "Detailed findings with specific remediation guidance follow.")
        story.append(Paragraph(summary, body_s))
        story.append(Spacer(1,0.3*cm))

        # AI ASSESSMENT SECTION
        story.append(HRFlowable(width="100%",thickness=1,color=LBLUE))
        story.append(Spacer(1,0.2*cm))

        has_real_ai = bool(ai_insights and ai_insights.get("summary"))

        if has_real_ai:
            # Real Claude assessment, passed through from the UI as-is — same
            # data shown in the "Claude AI Security Assessment" panel on screen.
            section_title = "Claude AI Security Assessment"
            ai_score = ai_insights.get("securityScore", overall)
            ai_summary = ai_insights.get("summary", "")
            str_items  = (ai_insights.get("strengths") or [])[:6] or ["No strengths reported"]
            risk_items = (ai_insights.get("risks") or [])[:6] or ["No risks reported"]
            rec_items  = (ai_insights.get("topRecommendations") or [])[:6] or ["No recommendations reported"]
        else:
            # Fallback: no real AI assessment was supplied (not run, or it failed)
            # — approximate from the pattern-matching results and say so plainly.
            section_title = "Pattern-Based Security Summary (AI assessment unavailable)"
            ai_score = overall
            strengths, risks, recs = [], [], []
            for cat, cd in sorted(results.items(), key=lambda x:x[1].get("pct",0), reverse=True):
                if cd.get("implemented",0)>0:
                    ctrl = next((r for r in cd.get("reqs",[]) if r.get("implemented") and r.get("finding")),None)
                    if ctrl:
                        strengths.append(cat+": "+ctrl["finding"].get("note","Detected")[:70])
                if cd.get("pct",0)<40:
                    gap = next((r for r in cd.get("reqs",[]) if not r.get("implemented")),None)
                    if gap:
                        risks.append(cat+": "+gap.get("requirement",""))
                        recs.append(gap.get("requirement",""))
            ai_summary = ("No Claude AI assessment was available for this export — this section is derived "
                "directly from pattern-matching coverage, not an independent review. The codebase shows "
                +str(total_found)+" ASVS 5.0 controls detected across "+str(cats_active)+" of 11 categories. "
                + ("Strong pattern coverage in "+", ".join(top_cats[:3])+". " if top_cats else ""))
            str_items = strengths[:4] if strengths else ["Pattern detection identified implemented controls"]
            risk_items = [r.split(": ",1)[-1][:85] if ": " in r else r[:85] for r in risks[:4]] or ["Review undetected categories"]
            rec_items  = [r.split("] ",1)[-1][:85] if "] " in r else r[:85] for r in recs[:4]]  or ["Improve low-coverage categories"]

        story.append(Paragraph(section_title, h1_s))

        ai_top = Table([[
            Paragraph("Security Analysis — ASVS Compliance & Maturity Analyzer", S("ah",fontSize=12,fontName="Helvetica-Bold",textColor=BLUE)),
            Paragraph("AI Score: "+str(ai_score)+"/100",
                S("as",fontSize=11,fontName="Helvetica-Bold",
                  textColor=GREEN if ai_score>=70 else AMBER if ai_score>=40 else RED,
                  alignment=TA_RIGHT))
        ]], colWidths=[12*cm,5*cm])
        ai_top.setStyle(TableStyle([
            ("BACKGROUND",(0,0),(-1,-1),LLBLUE),
            ("TOPPADDING",(0,0),(-1,-1),10),
            ("BOTTOMPADDING",(0,0),(-1,-1),10),
            ("LEFTPADDING",(0,0),(-1,-1),12),
            ("RIGHTPADDING",(0,0),(-1,-1),12),
        ]))
        story.append(ai_top)

        ai_sum_t = Table([[Paragraph(ai_summary, body_s)]], colWidths=[17*cm])
        ai_sum_t.setStyle(TableStyle([
            ("BACKGROUND",(0,0),(-1,-1),LLBLUE),
            ("LEFTPADDING",(0,0),(-1,-1),12),
            ("RIGHTPADDING",(0,0),(-1,-1),12),
            ("TOPPADDING",(0,0),(-1,-1),8),
            ("BOTTOMPADDING",(0,0),(-1,-1),8),
        ]))
        story.append(ai_sum_t)

        def ai_col(title, items, col, bg):
            rows = [[Paragraph(title, S("ct",fontSize=10,fontName="Helvetica-Bold",textColor=col))]]
            for item in (items[:6] if items else ["No items detected"]):
                rows.append([Paragraph("› "+str(item), S("ci",fontSize=8.5,fontName="Helvetica",textColor=DGRAY,leading=13,wordWrap="CJK"))])
            t = Table(rows, colWidths=[5.3*cm])
            t.setStyle(TableStyle([
                ("BACKGROUND",(0,0),(-1,-1),bg),
                ("TOPPADDING",(0,0),(-1,-1),6),
                ("BOTTOMPADDING",(0,0),(-1,-1),5),
                ("LEFTPADDING",(0,0),(-1,-1),8),
                ("RIGHTPADDING",(0,0),(-1,-1),8),
                ("LINEBELOW",(0,0),(0,0),1.5,col),
                ("VALIGN",(0,0),(-1,-1),"TOP"),
            ]))
            return t

        ai_cols = Table([[
            ai_col("Strengths", str_items, GREEN, colors.HexColor("#F0FFF0")),
            ai_col("Security Risks",     risk_items, RED,   colors.HexColor("#FFF5F5")),
            ai_col("Recommendations",    rec_items,  AMBER, colors.HexColor("#FFFDF0")),
        ]], colWidths=[5.6*cm,5.6*cm,5.6*cm])
        ai_cols.setStyle(TableStyle([
            ("LEFTPADDING",(0,0),(-1,-1),4),
            ("RIGHTPADDING",(0,0),(-1,-1),4),
            ("TOPPADDING",(0,0),(-1,-1),6),
            ("BOTTOMPADDING",(0,0),(-1,-1),6),
            ("VALIGN",(0,0),(-1,-1),"TOP"),
            ("ALIGN",(0,0),(-1,-1),"LEFT"),
        ]))
        story.append(ai_cols)
        story.append(Spacer(1,0.4*cm))

        # CATEGORY BREAKDOWN
        story.append(HRFlowable(width="100%",thickness=1,color=LBLUE))
        story.append(Spacer(1,0.2*cm))
        story.append(Paragraph("Category Breakdown", h1_s))

        ch = [
            Paragraph("<b>Security Category</b>", S("th",fontSize=9,fontName="Helvetica-Bold",textColor=WHITE,alignment=TA_CENTER)),
            Paragraph("<b>Found</b>",   S("th",fontSize=9,fontName="Helvetica-Bold",textColor=WHITE,alignment=TA_CENTER)),
            Paragraph("<b>Total</b>",   S("th",fontSize=9,fontName="Helvetica-Bold",textColor=WHITE,alignment=TA_CENTER)),
            Paragraph("<b>Coverage</b>",S("th",fontSize=9,fontName="Helvetica-Bold",textColor=WHITE,alignment=TA_CENTER)),
            Paragraph("<b>Maturity</b>",S("th",fontSize=9,fontName="Helvetica-Bold",textColor=WHITE,alignment=TA_CENTER)),
            Paragraph("<b>Top Finding / Gap</b>",S("th",fontSize=9,fontName="Helvetica-Bold",textColor=WHITE)),
        ]
        crow=[ch]
        cst=[
            ("BACKGROUND",(0,0),(-1,0),NAVY),
            ("GRID",(0,0),(-1,-1),0.5,MGRAY),
            ("TOPPADDING",(0,0),(-1,-1),5),
            ("BOTTOMPADDING",(0,0),(-1,-1),5),
            ("LEFTPADDING",(0,0),(-1,-1),6),
            ("ALIGN",(1,0),(4,-1),"CENTER"),
            ("VALIGN",(0,0),(-1,-1),"MIDDLE"),
        ]
        for i,(cat,cd) in enumerate(results.items(),1):
            imp=cd.get("implemented",0); tot=cd.get("total",0)
            pct, status = compute_weighted_score(cd.get("reqs", []))
            status = "Below L1" if status=="Below Level 1" else status
            sc=GREEN if status=="Level 2" else AMBER if status=="Level 1" else RED
            bg=LLBLUE if i%2==0 else WHITE
            top = next((r for r in cd.get("reqs",[]) if r.get("implemented") and r.get("finding")),None)
            gap = next((r for r in cd.get("reqs",[]) if not r.get("implemented")),None)
            if top and top.get("finding"):
                note=top["finding"].get("note","")[:55]; nc=GREEN
            elif gap:
                note="Gap: "+gap.get("requirement","")[:48]+"..."; nc=RED
            else:
                note="—"; nc=GRAY
            crow.append([
                Paragraph(cat,             S("cn",fontSize=9,fontName="Helvetica")),
                Paragraph(str(imp),        S("cv",fontSize=9,fontName="Helvetica",alignment=TA_CENTER)),
                Paragraph(str(tot),        S("cv",fontSize=9,fontName="Helvetica",alignment=TA_CENTER)),
                Paragraph(str(pct)+"%",    S("cv",fontSize=9,fontName="Helvetica-Bold",alignment=TA_CENTER)),
                Paragraph(status,          S("cs",fontSize=9,fontName="Helvetica-Bold",textColor=sc,alignment=TA_CENTER)),
                Paragraph(str(note).replace("&","&amp;").replace("<","&lt;").replace(">","&gt;"),S("fn",fontSize=8,fontName="Helvetica",textColor=nc)),
            ])
            cst.append(("BACKGROUND",(0,i),(-1,i),bg))

        crow.append([
            Paragraph("<b>TOTAL</b>",S("tot",fontSize=9,fontName="Helvetica-Bold")),
            Paragraph("<b>"+str(total_found)+"</b>",S("tot",fontSize=9,fontName="Helvetica-Bold",alignment=TA_CENTER)),
            Paragraph("<b>"+str(total_reqs)+"</b>", S("tot",fontSize=9,fontName="Helvetica-Bold",alignment=TA_CENTER)),
            Paragraph("<b>"+str(overall)+"%</b>",   S("tot",fontSize=9,fontName="Helvetica-Bold",alignment=TA_CENTER)),
            Paragraph("",body_s), Paragraph("",body_s),
        ])
        cst.append(("BACKGROUND",(0,len(crow)-1),(-1,len(crow)-1),LBLUE))
        cat_table=Table(crow,colWidths=[3.5*cm,1.5*cm,1.5*cm,2*cm,2*cm,6.5*cm])
        cat_table.setStyle(TableStyle(cst))
        story.append(cat_table)
        story.append(PageBreak())

        # DETAILED FINDINGS
        story.append(Paragraph("Detailed Findings per Category", h1_s))

        remed = {
            # Authentication
            "2.1.1":"Add length check: if len(password) < 12: raise ValidationError",
            "2.1.2":"Remove any max password length restriction from registration form",
            "2.1.3":"Never truncate passwords — hash and store full password string",
            "2.1.4":"Remove character type restrictions — accept Unicode and spaces",
            "2.1.5":"Add /account/change-password endpoint for authenticated users",
            "2.1.6":"Require request.form['current_password'] before accepting new password",
            "2.1.7":"Call pwnedpasswords API: GET https://api.pwnedpasswords.com/range/{hash}",
            "2.1.8":"Add zxcvbn.js to registration form — show strength meter to user",
            "2.1.9":"Delete all regex requiring uppercase, numbers, or symbols in password",
            "2.1.10":"Remove any forced password expiry or password history check logic",
            "2.1.11":"Remove JavaScript preventing paste — allow password manager autofill",
            "2.1.12":"Add <button onclick=toggle()> to show/hide password field content",
            "2.2.1":"Install Flask-Limiter: @limiter.limit('5 per minute') on /login",
            "2.2.2":"Replace SMS OTP with speakeasy TOTP authenticator app integration",
            "2.2.3":"Send email alert on password change and new device login events",
            "2.2.4":"Use TOTP or FIDO2 for high-risk operations per risk assessment",
            "2.2.5":"Implement pyotp TOTP: totp = pyotp.TOTP(secret); totp.verify(token)",
            "2.2.6":"Integrate Google Authenticator: pip install pyotp, generate base32 secret",
            "2.2.7":"Add WebAuthn support for FIDO2 hardware security key authentication",
            "2.3.1":"Generate temp passwords: secrets.token_hex(8), expire after 24 hours",
            "2.3.2":"Set must_change_password=True flag, force redirect on first login",
            "2.3.3":"Set 10 minute TTL on password reset tokens in database",
            "2.4.1":"Replace MD5/SHA1 with: hashed = bcrypt.hashpw(pwd.encode(), bcrypt.gensalt(12))",
            "2.4.2":"bcrypt auto-generates unique 32-bit salt — ensure not using custom salt",
            "2.4.3":"Set PBKDF2 iterations: hashlib.pbkdf2_hmac('sha256', pwd, salt, 100000)",
            "2.4.4":"Set bcrypt cost factor: bcrypt.gensalt(rounds=12) — minimum rounds=10",
            "2.4.5":"Install argon2-cffi: ph = PasswordHasher(time_cost=2, memory_cost=65536)",
            "2.5.1":"Send reset link via HTTPS: https://app.com/reset?token=<secure_token>",
            "2.5.2":"Delete all security question fields from database and registration form",
            "2.5.3":"Require email verification code as sole recovery mechanism",
            "2.5.4":"Search codebase for admin:admin root:root — remove all default credentials",
            "2.5.5":"On password change: db.execute('DELETE FROM sessions WHERE user_id=?', id)",
            "2.5.6":"Generate reset token: secrets.token_urlsafe(32), expire in 10 minutes",
            "2.5.7":"Require current password verification before allowing MFA device changes",
            "2.7.1":"Replace SMS: use pyotp.TOTP, generate QR code with qrcode library",
            "2.7.2":"Set OOB request expiry: expires_at = datetime.now() + timedelta(minutes=10)",
            "2.7.3":"Invalidate previous OOB token on new request: UPDATE tokens SET active=0",
            "2.10.1":"Replace static API keys with JWT: jwt.encode(payload, key, algorithm='RS256')",
            "2.10.2":"Audit all service accounts and rotate any default or weak passwords",
            "2.10.3":"Store service passwords in HashiCorp Vault: vault kv put secret/svc pass=X",
            "2.10.4":"Move secrets to .env: load_dotenv(); key = os.getenv('SECRET_KEY')",
            # Session Management
            "3.1.1":"Never append session token to URL — use Set-Cookie header only",
            "3.2.1":"After login: session.regenerate() or session.clear(); session['user']=id",
            "3.2.2":"Use Flask session with SECRET_KEY of 32+ random bytes from os.urandom",
            "3.2.3":"Store session in httpOnly cookie only — remove any localStorage session",
            "3.2.4":"Generate session ID: secrets.token_hex(32) via CSPRNG",
            "3.3.1":"On logout: session.clear(); response.delete_cookie('session')",
            "3.3.2":"Set PERMANENT_SESSION_LIFETIME = timedelta(hours=8) in Flask config",
            "3.3.3":"On password change: invalidate all other active sessions for this user",
            "3.3.4":"Add /account/sessions endpoint to list and terminate active sessions",
            "3.4.1":"Set-Cookie: session=X; Secure — add secure=True to all cookie calls",
            "3.4.2":"Set-Cookie: session=X; HttpOnly — add httponly=True to all cookies",
            "3.4.3":"Set-Cookie: session=X; SameSite=Strict — add samesite='Strict'",
            "3.4.4":"Set cookie path='/app' instead of path='/' for session cookies",
            "3.4.5":"Remove Domain attribute unless subdomain sharing is explicitly required",
            "3.5.1":"Implement OAuth2 client_credentials flow for service-to-service auth",
            "3.5.2":"Replace server sessions with: jwt.encode(payload, key, algorithm='RS256')",
            "3.5.3":"Use RS256: jwt.encode(payload, private_key, algorithm='RS256') only",
            "3.7.1":"Add password confirmation step before password change and account deletion",
            # Access Control
            "4.1.1":"Move all @login_required and permission checks to Flask route decorators",
            "4.1.2":"Validate user role claims from JWT before using in access decisions",
            "4.1.3":"Add check: if resource.user_id != current_user.id: return 403",
            "4.1.5":"Return generic 403 Forbidden — never expose internal permission logic",
            "4.2.1":"Check ownership: if db.get(id).owner != user.id: abort(403)",
            "4.2.2":"Install Flask-WTF: app.config['WTF_CSRF_ENABLED'] = True globally",
            "4.3.1":"Add @require_mfa decorator to all /admin and sensitive routes",
            "4.3.2":"Add to nginx: autoindex off; or Apache: Options -Indexes",
            "4.3.3":"Add re-authentication step for payments, deletion, and MFA changes",
            # Input Validation
            "5.1.1":"Check for duplicate params: if len(set(request.args.keys())) != len(list(request.args.keys())): abort(400)",
            "5.1.2":"Use explicit field allowlist: allowed = ['name','email']; data = {k:v for k,v in form.items() if k in allowed}",
            "5.1.3":"Use Marshmallow schema with strict field definitions and unknown=EXCLUDE",
            "5.1.4":"Validate with jsonschema: jsonschema.validate(instance=data, schema=schema)",
            "5.1.5":"Check redirect: if redirect_url not in ALLOWED_URLS: abort(400)",
            "5.2.1":"Install bleach: clean_html = bleach.clean(user_input, tags=ALLOWED_TAGS)",
            "5.2.2":"Apply regex allowlist: re.sub(r'[^a-zA-Z0-9 .,!?]', '', user_input)",
            "5.2.3":"Escape newlines in email: content.replace('\r\n','').replace('\n','')",
            "5.2.4":"Replace all eval() with json.loads() or ast.literal_eval()",
            "5.2.5":"Use Jinja2 with autoescaping: Environment(autoescape=True)",
            "5.2.6":"Validate URL: if urlparse(url).netloc not in ALLOWED_HOSTS: abort(400)",
            "5.2.7":"Strip SVG scripts: re.sub(r'<script[^>]*>.*?</script>', '', svg)",
            "5.2.8":"Sanitize markdown: bleach.clean(markdown.markdown(user_input))",
            "5.3.1":"Use Jinja2 autoescape=True and url_encode() for URL context output",
            "5.3.2":"Set Content-Type: text/html; charset=UTF-8 on all responses",
            "5.3.3":"Enable Jinja2 autoescaping: render_template() with autoescape=True",
            "5.3.4":"Use SQLAlchemy ORM: db.session.execute(text('SELECT * WHERE id=:id'), {'id':id})",
            "5.3.5":"Use db.escape(value) on any dynamic SQL that cannot be parameterized",
            "5.3.6":"Parse JSON strictly: json.loads() — never exec() or eval() on JSON",
            "5.3.7":"Use ldap3.utils.escape_filter_chars() on all LDAP query parameters",
            "5.3.8":"Replace os.system(cmd) with subprocess.run(['cmd','arg'], shell=False)",
            "5.3.9":"Validate file paths: if not path.startswith(ALLOWED_DIR): abort(403)",
            "5.3.10":"Set lxml parser: etree.parse(f, etree.XMLParser(resolve_entities=False))",
            "5.5.1":"Sign data: hmac.new(key, data, sha256).hexdigest() — verify before use",
            "5.5.2":"Replace yaml.load() with yaml.safe_load() throughout codebase",
            "5.5.3":"Replace pickle.loads() with json.loads() for all user-supplied data",
            "5.5.4":"Replace eval(jsonStr) with JSON.parse(jsonStr) in all JavaScript",
            # Cryptography
            "6.1.1":"Encrypt PII: encrypted = Fernet(key).encrypt(pii_value.encode())",
            "6.1.2":"Apply AES-256-GCM to all health data fields before database insert",
            "6.1.3":"Use column encryption for card numbers: encrypt before save, decrypt on read",
            "6.2.1":"Wrap crypto in try/except — return generic error never padding details",
            "6.2.2":"Use AES-256-GCM only: cipher = Cipher(algorithms.AES(key), modes.GCM(iv))",
            "6.2.3":"Generate fresh IV: iv = os.urandom(12) — use GCM mode for authentication",
            "6.2.4":"Store algorithm name in config.py — never hardcode cipher string in code",
            "6.2.5":"Replace hashlib.md5/sha1 with hashlib.sha256 or bcrypt for passwords",
            "6.2.6":"Generate new IV per operation: iv = os.urandom(12) — never reuse IV",
            "6.2.7":"Switch to AES-GCM: modes.GCM(iv) provides built-in authentication tag",
            "6.2.8":"Compare tokens: hmac.compare_digest(token_a, token_b) — not ==",
            "6.3.1":"Replace random.random() with secrets.token_hex(32) for all tokens",
            "6.3.2":"Replace uuid.uuid1() with uuid.uuid4() for all GUID generation",
            "6.3.3":"Use /dev/urandom and monitor entropy pool under high load conditions",
            "6.4.1":"Store keys in Vault: client.secrets.kv.v2.read_secret(path='keys/app')",
            "6.4.2":"Use Vault transit: vault.write('transit/encrypt/key', plaintext=base64(data))",
            # Logging
            "7.1.1":"Add log filter: logging.Filter that replaces password= with password=***",
            "7.1.2":"Mask PII in logs: email[:3]+'***@'+email.split('@')[1] before logging",
            "7.1.3":"Log auth: app.logger.info('LOGIN user=%s ip=%s result=%s', user, ip, result)",
            "7.1.4":"Include in all security logs: user_id, remote_addr, timestamp, action",
            "7.2.1":"Log auth decisions but filter out password and token field values",
            "7.2.2":"Log 403 denials: logger.warning('ACCESS_DENIED user=%s resource=%s', u, r)",
            "7.3.1":"Escape newlines: log_msg.replace('\n','\\n').replace('\r','\\r')",
            "7.3.2":"Sanitize log values before displaying in admin dashboard",
            "7.3.3":"Send logs to syslog or CloudWatch — use append-only log storage",
            "7.4.1":"@app.errorhandler(500): return jsonify(error='Internal error'), 500",
            "7.4.2":"Wrap DB/API calls: try: result = db.execute(q) except Exception as e: log(e)",
            "7.4.3":"Register: @app.errorhandler(Exception) to catch all unhandled exceptions",
            # Data Protection
            "8.1.1":"Set Cache-Control: no-store on API responses containing sensitive data",
            "8.1.2":"Invalidate cached sensitive data after each authorized access",
            "8.1.3":"Remove unused hidden fields from all HTML forms",
            "8.1.4":"Add rate limiting to all public endpoints: @limiter.limit('100/hour')",
            "8.2.1":"Add header: response.headers['Cache-Control'] = 'no-store, no-cache'",
            "8.2.2":"Move sensitive data from localStorage to httpOnly session cookie",
            "8.2.3":"On logout: localStorage.clear(); sessionStorage.clear(); clear DOM",
            "8.3.1":"Move API keys from ?api_key=X to Authorization: Bearer X header",
            "8.3.2":"Add GDPR endpoints: GET /account/export and DELETE /account/delete",
            "8.3.3":"Add consent checkbox before collecting email or personal information",
            "8.3.4":"Create data inventory document listing all PII fields and storage location",
            "8.3.5":"Log all PII access: logger.info('DATA_ACCESS user=%s field=%s', u, field)",
            "8.3.6":"Zero out sensitive bytes after use: ctypes.memset(buf, 0, len(buf))",
            "8.3.7":"Encrypt PII columns: cipher.encrypt(value.encode()) before db.save()",
            "8.3.8":"Add scheduled job: DELETE FROM users WHERE created_at < DATE_SUB(NOW(), INTERVAL 7 YEAR)",
            # Communication Security
            "9.1.1":"Add before_request: if not request.is_secure: return redirect(https_url, 301)",
            "9.1.2":"Set nginx ssl_ciphers to ECDHE-RSA-AES256-GCM-SHA384 only",
            "9.1.3":"Set: context.minimum_version = ssl.TLSVersion.TLSv1_2 in SSL context",
            "9.2.1":"Set verify=True in all requests.get/post — provide CA bundle path",
            "9.2.2":"Enable TLS on all internal database and microservice connections",
            "9.2.3":"Remove verify=False from all requests calls — never disable cert check",
            "9.2.4":"Add to nginx: ssl_stapling on; ssl_stapling_verify on; resolver 8.8.8.8",
            "9.2.5":"Log TLS errors: except ssl.SSLError as e: logger.error('TLS error: %s', e)",
            # Files
            "12.1.1":"Add: app.config['MAX_CONTENT_LENGTH'] = 10 * 1024 * 1024  # 10MB limit",
            "12.1.2":"Check: if len(zf.read(name)) > 100 * compressed_size: abort(400)",
            "12.2.1":"Use python-magic: if magic.from_buffer(data) != expected_mime: abort(400)",
            "12.2.2":"Verify magic bytes match claimed Content-Type before processing",
            "12.2.3":"Validate XML: lxml.etree.XMLSchema(xsd).validate(doc) before parsing",
            "12.3.1":"Sanitize filename: safe = os.path.basename(filename) before file ops",
            "12.3.2":"Check filename: if not re.match(r'^[a-zA-Z0-9._-]+$', name): abort(400)",
            "12.3.3":"Validate remote URL: if urlparse(url).netloc not in WHITELIST: abort(400)",
            "12.3.4":"Add header: Content-Disposition: attachment; filename=file.pdf",
            "12.3.5":"Never pass user filename to os.system() or subprocess directly",
            "12.4.1":"Set UPLOAD_FOLDER='/var/uploads' — outside Flask static directory",
            "12.4.2":"Scan with ClamAV: subprocess.run(['clamscan', filepath]) before use",
            "12.5.1":"Configure nginx to serve only .html .css .js .png extensions",
            "12.5.2":"Set X-Content-Type-Options: nosniff on all file download responses",
            "12.6.1":"Validate fetch URLs: if domain not in ALLOWED_DOMAINS: abort(400)",
            # API
            "13.1.1":"Use single JSON parser across all components — standardize on json module",
            "13.1.2":"Add OpenAPI schema validation to all endpoints using flask-smorest",
            "13.1.3":"Move ?api_key=X to Authorization: Bearer X in request headers",
            "13.1.4":"Check permissions at both route level @login_required and object level",
            "13.1.5":"Add: if request.content_type != 'application/json': abort(415)",
            "13.2.1":"Specify methods: @app.route('/api', methods=['GET','POST']) only",
            "13.2.2":"Validate body: jsonschema.validate(request.json, REQUEST_SCHEMA)",
            "13.2.3":"Enable Flask-WTF CSRF or add X-CSRF-Token header validation",
            "13.2.4":"Use Marshmallow: class Schema(Schema): name=fields.Str(required=True)",
            "13.2.5":"Check: if 'application/json' not in request.content_type: abort(415)",
            "13.2.6":"Ensure all API endpoints are served exclusively over HTTPS with TLS",
            "13.3.1":"Parse XML: defusedxml.ElementTree.parse(source) — never lxml directly",
            "13.3.2":"Add WS-Security signature header to all outbound SOAP messages",
            "13.4.1":"Install graphql-depth-limit: depthLimit(5) and complexity limits",
            "13.4.2":"Add permission check in every GraphQL resolver function",
            # Configuration
            "14.1.1":"Create separate dev.env staging.env and prod.env configuration files",
            "14.1.2":"Migrate server config to Terraform HCL or Ansible playbook YAML",
            "14.1.3":"Add SAST stage to GitHub Actions: uses: returntocorp/semgrep-action",
            "14.1.4":"Use RBAC for deployments — restrict prod access to senior engineers",
            "14.1.5":"Log all deployments: timestamp, deployer, git_sha to audit trail",
            "14.2.1":"Add to CI/CD: pip install safety && safety check -r requirements.txt",
            "14.2.2":"Run: pip-autoremove && remove all dev dependencies from production",
            "14.2.3":"Add integrity attribute: <script src=cdn.js integrity=sha384-XXX crossorigin=anonymous>",
            "14.2.4":"Remove any packages from GitHub URLs or unmaintained PyPI packages",
            "14.2.5":"Pin versions with hashes: pip install --require-hashes -r requirements.txt",
            "14.2.6":"Disable unused Flask extensions and comment out unused middleware",
            "14.3.1":"Add: @app.errorhandler(404) and @app.errorhandler(500) with custom pages",
            "14.3.2":"Set in production.env: FLASK_DEBUG=0 FLASK_ENV=production",
            "14.3.3":"Add to nginx: server_tokens off; and remove X-Powered-By Flask header",
            "14.4.1":"Add explicit Content-Type header to all Flask response objects",
            "14.4.2":"Add: response.headers['Content-Disposition'] = 'attachment' on downloads",
            "14.4.3":"Add after_request: response.headers[CSP] = default-src self — in Flask after_request hook",
            "14.4.4":"Add after_request: response.headers['X-Content-Type-Options'] = 'nosniff'",
            "14.4.5":"Add after_request: response.headers['Strict-Transport-Security'] = 'max-age=31536000; includeSubDomains'",
            "14.4.6":"Add after_request: response.headers['Referrer-Policy'] = 'no-referrer'",
            "14.4.7":"Add after_request: response.headers['X-Frame-Options'] = 'DENY'",
            "14.5.1":"Specify methods=['GET','POST'] in route decorators — block all others",
            "14.5.2":"Never use request.headers.get('Origin') for authentication decisions",
            "14.5.3":"Set CORS origins=['https://yourdomain.com'] — never origins=['*']",
            "14.5.4":"Validate all X-Forwarded-* and SSO headers before trusting them",
        }

        for cat,cd in results.items():
            reqs=cd.get("reqs",[]); imp=cd.get("implemented",0); tot=cd.get("total",0)
            pct, cat_level_banner = compute_weighted_score(reqs)
            cc=GREEN if cat_level_banner=="Level 2" else BLUE if cat_level_banner=="Level 1" else RED
            cbanner=Table([[
                Paragraph("<b>"+cat.replace("&","&amp;")+"</b>",S("cb",fontSize=11,fontName="Helvetica-Bold",textColor=WHITE)),
                Paragraph(str(imp)+"/"+str(tot)+"  ("+str(pct)+"%)",
                    S("cr",fontSize=10,fontName="Helvetica",textColor=WHITE,alignment=TA_RIGHT)),
            ]],colWidths=[12*cm,5*cm])
            cbanner.setStyle(TableStyle([
                ("BACKGROUND",(0,0),(-1,-1),cc),
                ("TOPPADDING",(0,0),(-1,-1),8),("BOTTOMPADDING",(0,0),(-1,-1),8),
                ("LEFTPADDING",(0,0),(0,-1),12),("RIGHTPADDING",(-1,0),(-1,-1),12),
                ("VALIGN",(0,0),(-1,-1),"MIDDLE"),
            ]))
            story.append(cbanner)
            story.append(Spacer(1,0.1*cm))

            rh=[
                Paragraph("<b>ID</b>",  S("rh",fontSize=8,fontName="Helvetica-Bold",textColor=WHITE,alignment=TA_CENTER)),
                Paragraph("<b>L</b>",   S("rh",fontSize=8,fontName="Helvetica-Bold",textColor=WHITE,alignment=TA_CENTER)),
                Paragraph("<b>Verification Requirement</b>",S("rh",fontSize=8,fontName="Helvetica-Bold",textColor=WHITE)),
                Paragraph("<b>Status</b>",S("rh",fontSize=8,fontName="Helvetica-Bold",textColor=WHITE,alignment=TA_CENTER)),
                Paragraph("<b>Finding / Remediation</b>",S("rh",fontSize=8,fontName="Helvetica-Bold",textColor=WHITE)),
            ]
            rrows=[rh]
            rst=[
                ("BACKGROUND",(0,0),(-1,0),NAVY),
                ("GRID",(0,0),(-1,-1),0.4,MGRAY),
                ("TOPPADDING",(0,0),(-1,-1),4),("BOTTOMPADDING",(0,0),(-1,-1),4),
                ("LEFTPADDING",(0,0),(-1,-1),5),
                ("ALIGN",(0,0),(1,-1),"CENTER"),("ALIGN",(3,0),(3,-1),"CENTER"),
                ("VALIGN",(0,0),(-1,-1),"MIDDLE"),
            ]
            for i,req in enumerate(reqs,1):
                done=req.get("implemented",False)
                f=req.get("finding",{}) or {}
                line_num = f.get("lineNumber") if done and f else None
                line_content = f.get("lineContent") if done and f else None
                conf = f.get("confidence","").upper() if done and f else ""
                if done and line_num and line_content:
                    note = f"Line {line_num}: {line_content[:70]}"
                elif done:
                    note = f.get("note","Control detected")[:80]
                else:
                    note = remed.get(req.get("id",""), "Implement: " + req.get("requirement","Review ASVS 5.0.0 spec for this requirement"))[:100]
                nc=GREEN if done else GRAY
                bg=LLBLUE if i%2==0 else WHITE
                rrows.append([
                    Paragraph(req.get("id",""),   S("ri",fontSize=8,fontName="Helvetica-Bold",textColor=BLUE,alignment=TA_CENTER)),
                    Paragraph(req.get("level",""),S("rl",fontSize=8,fontName="Helvetica",alignment=TA_CENTER)),
                    Paragraph((req.get("requirement","")[:120] + f" [Level {req.get('level','')} | {req.get('verificationMethod','SAST')}]").replace("&","&amp;").replace("<","&lt;").replace(">","&gt;"),S("rr",fontSize=8,fontName="Helvetica",leading=11)),
                    Paragraph("<b>Pass</b>" if done else "<b>Fail</b>",
                        S("rs",fontSize=8,fontName="Helvetica-Bold",textColor=GREEN if done else RED,alignment=TA_CENTER)),
                    Paragraph(str(note).replace("&","&amp;").replace("<","&lt;").replace(">","&gt;"),S("rf",fontSize=7.5,fontName="Helvetica",textColor=nc,leading=10)),
                ])
                rst.append(("BACKGROUND",(0,i),(-1,i),bg))
                rst.append(("BACKGROUND",(3,i),(3,i),LGREEN if done else LRED))

            rt=Table(rrows,colWidths=[1.4*cm,0.7*cm,6.5*cm,1.4*cm,6*cm])
            rt.setStyle(TableStyle(rst))
            story.append(rt)
            story.append(Spacer(1,0.4*cm))

        story.append(HRFlowable(width="100%",thickness=1.5,color=NAVY))
        story.append(Spacer(1,0.2*cm))
        story.append(Paragraph(
            "Generated by ASVS Compliance & Maturity Analyzer  ·  OWASP ASVS 5.0  ·  "+now+"  ·  "
            "Coverage percentages are security maturity indicators and do not constitute formal ASVS compliance certification.",
            footer_s))

        doc.build(story)
        buf.seek(0)
        return send_file(buf, mimetype="application/pdf",
            as_attachment=True, download_name="ASVS-Compliance-Maturity-Report.pdf")
    except Exception as e:
        import traceback
        print(traceback.format_exc())
        return jsonify({"error": str(e)}), 500


if __name__ == '__main__':
    init_db()
    print("ASVS Compliance & Maturity Analyzer backend running on port 3001...")
    app.run(port=3001, debug=False)
